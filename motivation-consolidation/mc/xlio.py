"""Єдиний інтерфейс читання .xls та .xlsx.

Файли-джерела приходять у двох форматах, тому весь інший код працює
з абстракцією Grid і не знає, який движок під ним.
Індексація всюди 1-based, як у самому Excel.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from functools import lru_cache
from typing import Any


def col_letter(idx: int) -> str:
    """1 -> A, 27 -> AA"""
    out = ""
    while idx:
        idx, rem = divmod(idx - 1, 26)
        out = chr(65 + rem) + out
    return out


def col_index(letter: str) -> int:
    idx = 0
    for ch in letter.upper():
        idx = idx * 26 + (ord(ch) - 64)
    return idx


class Grid:
    """Один аркуш; індексація 1-based, як у самому Excel."""

    name: str
    nrows: int
    ncols: int

    def value(self, row: int, col: int) -> Any:  # pragma: no cover - інтерфейс
        raise NotImplementedError

    # -- зручні похідні --------------------------------------------------

    def text(self, row: int, col: int) -> str:
        v = self.value(row, col)
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        return str(v).strip()

    def number(self, row: int, col: int) -> float | None:
        v = self.value(row, col)
        if isinstance(v, bool):
            return None
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            s = v.strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
            if re.fullmatch(r"-?\d+(\.\d+)?", s):
                return float(s)
        return None

    def row_texts(self, row: int, max_col: int | None = None) -> list[str]:
        limit = min(self.ncols, max_col or self.ncols)
        return [self.text(row, c) for c in range(1, limit + 1)]

    def find(self, pattern: str, max_row: int | None = None,
             max_col: int | None = None) -> list[tuple[int, int, str]]:
        """Усі клітинки з текстом, що матчить regex (без урахування регістру)."""
        rx = re.compile(pattern, re.I)
        hits = []
        rlimit = min(self.nrows, max_row or self.nrows)
        climit = min(self.ncols, max_col or self.ncols)
        for r in range(1, rlimit + 1):
            for c in range(1, climit + 1):
                t = self.text(r, c)
                if t and rx.search(t):
                    hits.append((r, c, t))
        return hits


class _XlsGrid(Grid):
    def __init__(self, sheet):
        self.name = sheet.name
        self.nrows = sheet.nrows
        self.ncols = sheet.ncols
        self._sh = sheet

    def value(self, row: int, col: int):
        if 1 <= row <= self.nrows and 1 <= col <= self.ncols:
            v = self._sh.cell_value(row - 1, col - 1)
            return None if v == "" else v
        return None


class _XlsxGrid(Grid):
    """Аркуш xlsx, вичитаний потоково і один раз.

    Звичайний режим openpyxl будує об'єкт на кожну клітинку — на книзі
    Алекс ком (48 аркушів) це коштує близько пів хвилини, а в браузері
    втричі більше. Потокове читання дає ті самі значення за частки
    секунди, тому аркуш матеріалізується цілком при першому зверненні.
    """

    def __init__(self, ws):
        self.name = ws.title
        self._ws = ws
        self._rows: dict[int, tuple] | None = None
        self._nrows = 0
        self._ncols = 0

    def _materialise(self) -> None:
        if self._rows is not None:
            return
        rows: dict[int, tuple] = {}
        widest = 0
        for index, values in enumerate(
                self._ws.iter_rows(values_only=True), start=1):
            if any(v is not None for v in values):
                rows[index] = values
                widest = max(widest, len(values))
        self._rows = rows
        self._nrows = max(rows) if rows else 0
        self._ncols = widest

    @property
    def nrows(self) -> int:
        self._materialise()
        return self._nrows

    @property
    def ncols(self) -> int:
        self._materialise()
        return self._ncols

    def value(self, row: int, col: int):
        self._materialise()
        values = self._rows.get(row)
        if values is None or col < 1 or col > len(values):
            return None
        return values[col - 1]


@dataclass
class Workbook:
    path: str
    sheets: list[Grid]

    def names(self) -> list[str]:
        return [s.name for s in self.sheets]

    def by_name(self, name: str) -> Grid | None:
        for s in self.sheets:
            if s.name == name:
                return s
        return None

    def index_of(self, name: str) -> int:
        for i, s in enumerate(self.sheets):
            if s.name == name:
                return i
        return -1


@lru_cache(maxsize=64)
def load(path: str) -> Workbook:
    """Читає книгу цілком. Кешується — один файл відкривається раз за прогін."""
    low = path.lower()
    if low.endswith(".xls"):
        import xlrd

        bk = xlrd.open_workbook(path, on_demand=False)
        return Workbook(path, [_XlsGrid(sh) for sh in bk.sheets()])

    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    return Workbook(path, [_XlsxGrid(wb[name]) for name in wb.sheetnames])
