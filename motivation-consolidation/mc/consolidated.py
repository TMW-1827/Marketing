"""Читання і оновлення зведеної таблиці.

Розкладка зведеної: місяці йдуть блоками колонок зліва направо, дата
місяця стоїть у рядку 2 над першою колонкою блоку. На аркуші три
незалежні таблиці — за продажами, за АКБ і за сумою вторинних продаж, —
у кожної власний рядок заголовків і власний рядок «Сума».

Закритий місяць займає 8 колонок, поточний (ще без факту) — 5.
Тому щомісяця відбувається дві дії: попередній блок дорощується до
восьми колонок, і праворуч додається новий блок на п'ять.
"""

from __future__ import annotations

import datetime as dt
import re
from copy import copy
from dataclasses import dataclass, field

from .months import Period
from .xlio import col_letter

SHEET = "Акції та мотивації ТК"
HEADER_MARK = "Дистрибютор"
TOTAL_MARK = "Сума"

# роль колонки -> як вона підписана в рядку заголовків таблиці
ROLE_PATTERNS = [
    ("bonus_plan", re.compile(r"^бонус\s+план", re.I)),
    ("bonus_fact", re.compile(r"^бонус\s+факт", re.I)),
    ("growth_plan", re.compile(r"^план\s+приріст", re.I)),
    ("growth_fact", re.compile(r"^факт\s+приріст", re.I)),
    ("compensation", re.compile(r"^форма\s+компенсації", re.I)),
    ("cost_fact", re.compile(r"^затратність.*факт\s*$", re.I)),
    ("cost_plan", re.compile(r"^затратність.*план\s*$", re.I)),
    ("promo_terms", re.compile(r"^умови\s+акцій", re.I)),
    ("plan", re.compile(r"^план\s+(продаж|акб|сума)", re.I)),
    ("fact", re.compile(r"^факт\s+(продаж|акб|сума)", re.I)),
]

METRIC_BY_HEADER = [
    ("akb", re.compile(r"^план\s+акб", re.I)),
    ("revenue", re.compile(r"^план\s+сума", re.I)),
    ("sales", re.compile(r"^план\s+продаж", re.I)),
]

CLOSED_ROLES = ["plan", "fact", "bonus_plan", "bonus_fact",
                "growth_plan", "growth_fact", "compensation", "cost_fact"]
OPEN_ROLES = ["plan", "bonus_plan", "growth_plan", "cost_plan", "promo_terms"]

# підписи колонок для нового блоку; {m} — номер місяця
LABELS = {
    "sales": {
        "plan": "План продаж {m}", "fact": "Факт продаж {m}",
        "bonus_plan": "Бонус план {m} грн", "bonus_fact": "Бонус факт {m}, грн",
        "growth_plan": "План приріст до минулого місяця, %",
        "growth_fact": "Факт приріст до минулого місяця, %",
        "compensation": "Форма компенсації",
        "cost_fact": "Затратність на мотивації з 1 пл, грн, факт",
        "cost_plan": "Затратність на мотивації з 1 пл, грн, план",
        "promo_terms": "Умови акцій {m}",
    },
    "akb": {
        "plan": "План АКБ {m}", "fact": "Факт АКБ {m}",
        "bonus_plan": "Бонус план {m} грн", "bonus_fact": "Бонус факт {m}, грн",
        "growth_plan": "План приріст АКБ до минулого місяця, %",
        "growth_fact": "Факт приріст до минулого місяця, %",
        "compensation": "Форма компенсації",
        "cost_fact": "Затратність на мотивації за 1 ТТ, грн, факт",
        "cost_plan": "Затратність на мотивації за 1 ТТ, грн, план",
        "promo_terms": "Умови акцій {m}",
    },
    "revenue": {
        "plan": "План сума вториних продаж {m} грн",
        "fact": "Факт сума вторинних продаж {m} грн",
        "bonus_plan": "Бонус план {m} грн", "bonus_fact": "Бонус факт {m}, грн",
        "growth_plan": "План приріст суми до минулого місяця, %",
        "growth_fact": "Факт приріст до минулого місяця, %",
        "compensation": "Форма компенсації",
        "cost_fact": "Затратність на мотивації з суми ВП, %, факт",
        "cost_plan": "Затратність на мотивації з суми ВП, %, план",
        "promo_terms": "Умови акцій {m}",
    },
}


@dataclass
class Table:
    metric: str
    header_row: int
    total_row: int
    rows: dict[str, int] = field(default_factory=dict)   # назва -> рядок

    @property
    def first_row(self) -> int:
        return self.header_row + 1

    @property
    def last_row(self) -> int:
        return self.total_row - 1


@dataclass
class MonthBlock:
    period: Period
    start: int
    width: int
    # роль -> колонка, окремо для кожної таблиці (заголовки в них свої)
    roles: dict[int, dict[str, int]] = field(default_factory=dict)

    @property
    def closed(self) -> bool:
        return self.width >= 8


class Consolidated:
    def __init__(self, path: str):
        import openpyxl

        self.path = path
        self.wb = openpyxl.load_workbook(path, data_only=False)
        self.ws = self.wb[SHEET]
        self.tables = self._read_tables()
        self.blocks = self._read_blocks()
        # значення, які не мають колонки в закритому місяці і тому
        # губляться при закритті — про них треба сказати вголос
        self.dropped: list[str] = []
        # зауваження до стану зведеної, які варто побачити після запису
        self.notes: list[str] = []
        # оформлення колонок, які бувають лише у відкритому місяці
        # («затратність план», «умови акцій»): при закритті ці ролі
        # зникають, тож зразок для нового місяця треба зберегти заздалегідь
        self._open_styles: dict[tuple[int, str], dict] = {}

    # -- читання ---------------------------------------------------------

    def _read_tables(self) -> list[Table]:
        ws = self.ws
        tables: list[Table] = []
        header_rows = [r for r in range(1, ws.max_row + 1)
                       if str(ws.cell(row=r, column=2).value or "").strip() == HEADER_MARK]
        for hr in header_rows:
            total_row = None
            for r in range(hr + 1, ws.max_row + 1):
                if str(ws.cell(row=r, column=2).value or "").strip() == TOTAL_MARK:
                    total_row = r
                    break
            if not total_row:
                continue
            metric = self._metric_of(hr)
            table = Table(metric=metric, header_row=hr, total_row=total_row)
            for r in range(hr + 1, total_row):
                name = str(ws.cell(row=r, column=2).value or "").strip()
                if name:
                    table.rows[name] = r
            tables.append(table)
        return tables

    def _metric_of(self, header_row: int) -> str:
        """Метрика таблиці — за найсвіжішим місяцем.

        Читати зліва не можна: у 2023–2024 таблиця продажів вимірювалась
        в АКБ і лише згодом перейшла на пляшки, тож найлівіші заголовки
        показали б застарілу метрику.
        """
        ws = self.ws
        for c in range(ws.max_column, 2, -1):
            t = str(ws.cell(row=header_row, column=c).value or "").strip()
            for metric, rx in METRIC_BY_HEADER:
                if rx.match(t):
                    return metric
        return "sales"

    def _read_blocks(self) -> list[MonthBlock]:
        ws = self.ws
        starts: list[tuple[int, Period]] = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=2, column=c).value
            if isinstance(v, dt.datetime):
                starts.append((c, Period(v.year, v.month)))
            elif isinstance(v, dt.date):
                starts.append((c, Period(v.year, v.month)))

        header_rows = [t.header_row for t in self.tables]

        def has_header(col: int) -> bool:
            return any(str(ws.cell(row=hr, column=col).value or "").strip()
                       for hr in header_rows)

        blocks = []
        for i, (col, period) in enumerate(starts):
            end = starts[i + 1][0] - 1 if i + 1 < len(starts) else ws.max_column
            while end > col and not has_header(end):
                end -= 1        # праворуч від останнього блоку бувають порожні колонки
            block = MonthBlock(period=period, start=col, width=end - col + 1)
            for table in self.tables:
                roles = {}
                for c in range(col, end + 1):
                    t = str(ws.cell(row=table.header_row, column=c).value or "").strip()
                    for role, rx in ROLE_PATTERNS:
                        if rx.match(t):
                            roles.setdefault(role, c)
                            break
                block.roles[table.header_row] = roles
            blocks.append(block)
        return blocks

    # -- доступ ----------------------------------------------------------

    def block(self, period: Period) -> MonthBlock | None:
        for b in self.blocks:
            if b.period == period:
                return b
        return None

    def table_for(self, metric: str) -> Table | None:
        for t in self.tables:
            if t.metric == metric:
                return t
        return None

    def read(self, period: Period, metric: str, distributor: str) -> dict:
        block = self.block(period)
        table = self.table_for(metric)
        if not block or not table or distributor not in table.rows:
            return {}
        row = table.rows[distributor]
        roles = block.roles.get(table.header_row, {})
        out = {}
        for role, col in roles.items():
            out[role] = self.ws.cell(row=row, column=col).value
        return out

    def names(self, metric: str) -> list[str]:
        table = self.table_for(metric)
        return list(table.rows) if table else []

    # -- запис -----------------------------------------------------------

    def ensure_month(self, period: Period) -> MonthBlock:
        """Готує колонки під місяць: закриває попередній, відкриває новий."""
        existing = self.block(period)
        if existing:
            return existing

        last = self.blocks[-1]
        if period <= last.period:
            raise ValueError(
                f"місяць {period} не новіший за останній у зведеній ({last.period})")
        if not last.closed:
            self._close_block(last)
        start = last.start + last.width
        return self._open_block(period, start)

    def _close_block(self, block: MonthBlock) -> None:
        """Перебудовує відкритий блок (5 колонок) у закритий (8)."""
        ws = self.ws
        saved: dict[int, dict[str, object]] = {}
        for table in self.tables:
            roles = block.roles.get(table.header_row, {})
            vals = {}
            for role, col in roles.items():
                vals[role] = {r: ws.cell(row=r, column=col).value
                              for r in range(table.first_row, table.total_row + 1)}
            saved[table.header_row] = vals

        for table in self.tables:
            for role, values in saved[table.header_row].items():
                if role in CLOSED_ROLES:
                    continue
                filled = sum(1 for v in values.values()
                             if v not in (None, ""))
                if filled:
                    self.dropped.append(
                        f"{block.period}: «{LABELS[table.metric][role].format(m=block.period.month)}» "
                        f"({filled} знач.) — у закритому місяці такої колонки немає")

        self._remember_open_styles(block)
        if not self._has_facts(block):
            self.notes.append(
                f"{block.period.label}: місяць закрито без факту. Приріст "
                f"наступного місяця рахується від факту цього, тож поки він "
                f"порожній, у тих клітинках буде помилка ділення")

        block.width = len(CLOSED_ROLES)
        for table in self.tables:
            roles = {}
            for i, role in enumerate(CLOSED_ROLES):
                col = block.start + i
                roles[role] = col
                label = LABELS[table.metric][role].format(m=f"{block.period.month:02d}")
                ws.cell(row=table.header_row, column=col).value = label
                keep = saved[table.header_row].get(role, {})
                for r in range(table.first_row, table.total_row + 1):
                    ws.cell(row=r, column=col).value = keep.get(r)
            block.roles[table.header_row] = roles
        self._restyle(block)
        self._write_month_header(block)

    def _remember_open_styles(self, block: MonthBlock) -> None:
        """Знімає оформлення ролей, яких у закритому місяці не буде."""
        ws = self.ws
        for table in self.tables:
            roles = block.roles.get(table.header_row, {})
            for role, col in roles.items():
                if role in CLOSED_ROLES:
                    continue
                rows = [table.header_row] + list(
                    range(table.first_row, table.total_row + 1))
                letter = col_letter(col)
                dims = ws.column_dimensions
                self._open_styles[(table.header_row, role)] = {
                    "rows": {r: copy(ws.cell(row=r, column=col)._style)
                             for r in rows},
                    "width": dims[letter].width if letter in dims else None,
                }

    def has_facts(self, period: Period) -> bool:
        """Чи заповнений факт за місяць — від нього рахується приріст наступного."""
        block = self.block(period)
        return bool(block) and self._has_facts(block)

    def _has_facts(self, block: MonthBlock) -> bool:
        for table in self.tables:
            col = block.roles.get(table.header_row, {}).get("fact")
            if not col:
                continue
            for r in range(table.first_row, table.last_row + 1):
                if self.ws.cell(row=r, column=col).value not in (None, ""):
                    return True
        return False

    def _open_block(self, period: Period, start: int) -> MonthBlock:
        ws = self.ws
        block = MonthBlock(period=period, start=start, width=len(OPEN_ROLES))
        for table in self.tables:
            roles = {}
            for i, role in enumerate(OPEN_ROLES):
                col = start + i
                roles[role] = col
                ws.cell(row=table.header_row, column=col).value = \
                    LABELS[table.metric][role].format(m=f"{period.month:02d}")
            block.roles[table.header_row] = roles
        self.blocks.append(block)
        self._restyle(block)
        self._write_month_header(block)
        return block

    # -- оформлення ------------------------------------------------------

    def _restyle(self, block: MonthBlock) -> None:
        """Переносить оформлення з відповідної колонки попереднього місяця.

        При закритті місяця колонки міняють призначення: там, де був
        плановий бонус, стає факт, а на місці «Умови акцій» — приріст.
        Формати при цьому лишаються від старої ролі, і бонус показувався б
        відсотком. Тому оформлення кожної колонки береться не з того, що
        було на цьому місці, а з колонки тієї самої ролі в попередньому
        місяці — там воно вже правильне.
        """
        ws = self.ws
        for table in self.tables:
            for role, col in block.roles.get(table.header_row, {}).items():
                saved = self._open_styles.get((table.header_row, role))
                if saved:
                    for r, style in saved["rows"].items():
                        ws.cell(row=r, column=col)._style = copy(style)
                    if saved["width"]:
                        ws.column_dimensions[col_letter(col)].width = saved["width"]
                    continue
                src = self._style_source(block, table, role)
                if not src:
                    continue
                rows = [table.header_row] + list(
                    range(table.first_row, table.total_row + 1))
                for r in rows:
                    ws.cell(row=r, column=col)._style = copy(
                        ws.cell(row=r, column=src)._style)
                self._copy_width(src, col)
                self._fix_text_column(role, col, table)

        first_src = self._style_source(block, self.tables[0], "plan")
        if first_src:
            ws.cell(row=2, column=block.start)._style = copy(
                ws.cell(row=2, column=first_src)._style)
        self._frame_block(block)

    def _frame_block(self, block: MonthBlock) -> None:
        """Відбиває місяць від сусіднього товстішою лінією.

        Такий роздільник стоїть у кінці кожного місяця, а не лише
        останнього, тому сусідні блоки не чіпаються — новому просто
        ставиться така сама межа.
        """
        edge = block.start + block.width - 1
        for table in self.tables:
            for r in [table.header_row] + list(
                    range(table.first_row, table.total_row + 1)):
                self._set_right_border(r, edge, "medium")

    def _set_right_border(self, row: int, col: int, style: str) -> None:
        from openpyxl.styles import Border, Side

        cell = self.ws.cell(row=row, column=col)
        b = cell.border
        cell.border = Border(left=b.left, top=b.top, bottom=b.bottom,
                             right=Side(style=style, color=b.right.color),
                             diagonal=b.diagonal, diagonalUp=b.diagonalUp,
                             diagonalDown=b.diagonalDown)

    def _fix_text_column(self, role: str, col: int, table: Table) -> None:
        """«Умови акцій» — вільний текст, а не число.

        Зразок для неї беремо з «Форми компенсації», а там стоїть
        відсотковий формат; для текстової колонки його треба зняти,
        інакше набране в ній число покажеться відсотком.
        """
        if role != "promo_terms":
            return
        from openpyxl.styles import Alignment

        for r in range(table.header_row, table.total_row + 1):
            cell = self.ws.cell(row=r, column=col)
            cell.number_format = "General"
            if r > table.header_row:
                cell.alignment = Alignment(
                    horizontal="left", vertical=cell.alignment.vertical,
                    wrap_text=True)

    # роль -> чим її замінити, якщо такої в попередніх місяцях не знайшлось
    STYLE_FALLBACK = {"cost_plan": "cost_fact", "promo_terms": "compensation",
                      "fact": "plan", "bonus_fact": "bonus_plan",
                      "growth_fact": "growth_plan"}

    def _style_source(self, block: MonthBlock, table: Table,
                      role: str) -> int | None:
        """Колонка тієї самої ролі в найближчому попередньому місяці."""
        earlier = sorted([b for b in self.blocks
                          if b.start < block.start and b.period < block.period],
                         key=lambda b: b.start, reverse=True)
        for wanted in (role, self.STYLE_FALLBACK.get(role)):
            if not wanted:
                continue
            for other in earlier:
                col = other.roles.get(table.header_row, {}).get(wanted)
                if col:
                    return col
        return None

    def _copy_width(self, src: int, dst: int) -> None:
        dims = self.ws.column_dimensions
        src_letter, dst_letter = col_letter(src), col_letter(dst)
        if src_letter in dims and dims[src_letter].width:
            dims[dst_letter].width = dims[src_letter].width

    def _write_month_header(self, block: MonthBlock) -> None:
        # шапка місяця обʼєднана на всю ширину блоку; при закритті блок
        # розширюється, тож старе обʼєднання треба перескласти
        end = block.start + block.width - 1
        for rng in list(self.ws.merged_cells.ranges):
            if rng.min_row == 2 and rng.min_col == block.start:
                self.ws.unmerge_cells(str(rng))
        if end > block.start:
            self.ws.merge_cells(start_row=2, start_column=block.start,
                                end_row=2, end_column=end)
        cell = self.ws.cell(row=2, column=block.start)
        cell.value = dt.datetime(block.period.year, block.period.month, 1)
        cell.number_format = "mmm-yy"

    def write(self, period: Period, metric: str, distributor: str,
              values: dict) -> list[str]:
        """Записує значення дистриб'ютора. Повертає перелік заповнених клітинок."""
        block = self.block(period)
        table = self.table_for(metric)
        if not block or not table:
            raise ValueError(f"немає блоку {period} або таблиці {metric}")
        if distributor not in table.rows:
            raise ValueError(f"у таблиці «{metric}» немає рядка «{distributor}»")

        row = table.rows[distributor]
        roles = block.roles[table.header_row]
        touched = []
        for role, value in values.items():
            col = roles.get(role)
            if col is None or value is None:
                continue
            self.ws.cell(row=row, column=col).value = value
            touched.append(f"{col_letter(col)}{row}")
        self._write_derived(block, table, row)
        return touched

    def _write_derived(self, block: MonthBlock, table: Table, row: int) -> None:
        """Проставляє формули приросту і затратності — так само, як у файлі."""
        ws = self.ws
        roles = block.roles[table.header_row]
        prev = self._prev_block(block)
        prev_fact = None
        if prev:
            prev_fact = prev.roles.get(table.header_row, {}).get("fact")

        def ref(col: int) -> str:
            return f"{col_letter(col)}{row}"

        if prev_fact and roles.get("plan") and roles.get("growth_plan"):
            ws.cell(row=row, column=roles["growth_plan"]).value = \
                f"={ref(roles['plan'])}/{ref(prev_fact)}-100%"
        if prev_fact and roles.get("fact") and roles.get("growth_fact"):
            ws.cell(row=row, column=roles["growth_fact"]).value = \
                f"={ref(roles['fact'])}/{ref(prev_fact)}-100%"
        if roles.get("cost_fact") and roles.get("bonus_fact") and roles.get("fact"):
            ws.cell(row=row, column=roles["cost_fact"]).value = \
                f"={ref(roles['bonus_fact'])}/{ref(roles['fact'])}"
        if roles.get("cost_plan") and roles.get("bonus_plan") and roles.get("plan"):
            ws.cell(row=row, column=roles["cost_plan"]).value = \
                f"={ref(roles['bonus_plan'])}/{ref(roles['plan'])}"

    def _prev_block(self, block: MonthBlock) -> MonthBlock | None:
        target = block.period.prev
        for b in self.blocks:
            if b.period == target:
                return b
        return None

    def refresh_totals(self, period: Period | None = None) -> list[str]:
        """Перебудовує рядки «Сума» під фактичні межі кожної таблиці.

        У файлі частина діапазонів заповзає в сусідню таблицю
        (наприклад SUM(C13:C21) у підсумку за АКБ), тож перерахунок
        заразом лагодить успадковані помилки.
        """
        ws = self.ws
        fixed = []
        blocks = [self.block(period)] if period else self.blocks
        for block in [b for b in blocks if b]:
            for table in self.tables:
                roles = block.roles.get(table.header_row, {})
                for role in ("plan", "fact", "bonus_plan", "bonus_fact"):
                    col = roles.get(role)
                    if not col:
                        continue
                    letter = col_letter(col)
                    want = f"=SUM({letter}{table.first_row}:{letter}{table.last_row})"
                    cell = ws.cell(row=table.total_row, column=col)
                    if cell.value != want:
                        if cell.value is not None:
                            fixed.append(
                                f"{letter}{table.total_row}: {cell.value} -> {want}")
                        cell.value = want
        return fixed

    def save(self, path: str | None = None) -> str:
        target = path or self.path
        self.wb.save(target)
        return target
