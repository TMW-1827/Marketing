"""Тести на синтетичних книгах — відтворюють ті розкладки джерел,
на яких раніше ламався розбір."""

import os
import sys
import tempfile

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from mc import config as cfgmod                      # noqa: E402
from mc.blocks import extract_blocks                  # noqa: E402
from mc.consolidated import Consolidated              # noqa: E402
from mc.discover import scan_file                     # noqa: E402
from mc.months import Period, find_year, parse_period  # noqa: E402
from mc.promo import extract_promo                    # noqa: E402
from mc.xlio import load                              # noqa: E402


def write(rows: dict, name: str = "лютий 2026 факт") -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name
    for coord, value in rows.items():
        ws[coord] = value
    path = os.path.join(tempfile.mkdtemp(), "джерело.xlsx")
    wb.save(path)
    return path


def sheet_of(path: str):
    load.cache_clear()
    return load(path).sheets[0]


# -- розбір періодів --------------------------------------------------------

def test_year_after_cyrillic_letter():
    # «2025р.» — кирилична «р» ламала межу слова
    assert find_year("Факт мотивації на листопад 2025р.") == 2025


def test_two_digit_year():
    assert parse_period("ФАКТ акцій і мотивацій червень 26") == Period(2026, 6)


def test_period_ordering():
    assert Period(2026, 1).prev == Period(2025, 12)
    assert Period(2025, 12).shift(2) == Period(2026, 2)


# -- відповідність назв -----------------------------------------------------

def test_longest_match_wins():
    cfg = cfgmod.load()
    assert cfg.resolve("Дистрибутор - Буковина Трек ІФ").consolidated \
        == "Буковина Трек ІФ"
    assert cfg.resolve("Дистрибютор - Буковина Трек").consolidated \
        == "Буковина Трек Чернівці"
    assert cfg.resolve("Дистрибютор Омела Маркет.").consolidated == "Омела Маркет"
    assert cfg.resolve("Дистрибютор ТОВ Омела.").consolidated == "Омела"


# -- заголовки аркушів ------------------------------------------------------

def test_header_copied_from_previous_month_is_corrected():
    """Заголовок липневого аркуша скопійовано з червневого."""
    wb = openpyxl.Workbook()
    first = wb.active
    first.title = "план червень 2026"
    first["B4"] = "План акцій і мотивацій на червень 2026. ТМ Трускавецька."
    first["B5"] = "Дистрибютор Тест"
    second = wb.create_sheet("план липень 2026")
    second["B4"] = "План акцій і мотивацій на червень 2026. ТМ Трускавецька."
    path = os.path.join(tempfile.mkdtemp(), "копіпаст.xlsx")
    wb.save(path)

    load.cache_clear()
    sf = scan_file(path)
    assert sf.find(Period(2026, 7), "план") is not None
    warned = [w for s in sf.sheets for w in s.warnings if "назви аркуша" in w]
    assert warned, "має бути попередження про заголовок"


# -- витяг блоку ------------------------------------------------------------

BASE = {
    "B4": "Факт акцій і мотивацій на лютий 2026. ТМ Трускавецька.",
    "B5": "Дистрибютор Тест",
    "B7": "Блок 1. Погоджувальна частина",
    "B9": "ТП/СВ", "C9": "ЗАДАЧА 1. ТМ Трускавецька (пляшки)",
    "F9": "бонус максимум",
    "C10": "план,пл", "D10": "факт, пл", "E10": "%",
}


def test_totals_and_bonus():
    rows = dict(BASE)
    rows.update({
        "B11": "ТП Перший", "C11": 1000, "D11": 1100, "F11": 500,
        "B12": "ТП Другий", "C12": 1000, "D12": 900, "F12": 500,
        "B13": "ТП Третій", "C13": 2000, "D13": 2200, "F13": 700,
        "B14": "Всього:", "C14": 4000, "D14": 4200,
        "E15": "Сума", "F15": 1700,
    })
    block = extract_blocks(sheet_of(write(rows)))[0]
    assert block.metric == "sales" and block.unit == "пл"
    assert block.plan == 4000 and block.fact == 4200
    assert block.bonus_total == 1700
    # два однакові плани поспіль не роблять другий рядок підсумковим
    assert len(block.leaves) == 3


def test_supervisor_with_own_territory_counts_in_total():
    """СВ має власні продажі, тож підсумок = рядки ТП + рядок СВ."""
    rows = dict(BASE)
    rows.update({
        "B11": "ТП Перший", "C11": 1000, "D11": 800, "F11": 500,
        "B12": "ТП Другий", "C12": 2000, "D12": 1600, "F12": 500,
        "B13": "СВ Старший", "C13": 500, "D13": 400, "F13": 300,
        "B14": "НТО Головний", "C14": 3500, "D14": 2800,
        "E15": "Сума", "F15": 1300,
    })
    block = extract_blocks(sheet_of(write(rows)))[0]
    assert block.plan == 3500
    assert sum(i.plan for i in block.leaves) == 3000


def test_block_without_bonuses_is_not_motivated():
    """Немає сум мотивації — блок не потрапляє у зведену."""
    rows = dict(BASE)
    rows.update({
        "B11": "ЕК Перший", "C11": 1000, "D11": 1100, "F11": 0,
        "B12": "ЕК Другий", "C12": 1000, "D12": 900, "F12": 0,
        "B13": "Всього:", "C13": 2000, "D13": 2000, "F13": 0,
    })
    block = extract_blocks(sheet_of(write(rows)))[0]
    assert block.motivated is False


def test_motivation_rate_column_is_not_mistaken_for_plan():
    """«мотивація за виконання плану» містить слово «план», але це бонус."""
    rows = {
        "B4": "Факт акцій і мотивацій на лютий 2026. ТМ Трускавецька.",
        "B5": "Дистрибютор Тест",
        "B9": "№", "C9": "ПІБ", "D9": "План 02", "E9": "ФАКТ 02",
        "F9": "ВИКОНАННЯ 02, %", "G9": "мотивація за виконання плану",
        "H9": "бонус , грн.",
        "B10": 1, "C10": "Перший", "D10": 1000, "E10": 1200,
        "G10": "ТП отримує бонус 1000 грн за виконання плану на 100 %",
        "H10": 1000,
        "B11": 2, "C11": "Другий", "D11": 2000, "E11": 1800, "H11": 500,
        "C12": "СВ Старший", "D12": 3000, "E12": 3000, "H12": 500,
        "D13": 3000, "H13": 2000,
        "B15": "Задачі на лютий: план продажів - 3 000 пл.",
    }
    blocks = extract_blocks(sheet_of(write(rows)))
    assert len(blocks) == 1, "має бути один блок, а не два"
    block = blocks[0]
    assert block.cols["plan"] == 4          # колонка D
    assert block.plan == 3000
    assert block.bonus_total == 2000        # підсумок під таблицею
    assert block.unit == "пл"               # одиницю взято з опису задачі


# -- акції ------------------------------------------------------------------

PROMO = {
    "B4": "Факт акцій і мотивацій на лютий 2026. ТМ Трускавецька.",
    "B5": "Дистрибютор Тест",
    "B7": "Блок 2. Акції",
}


def test_promo_balance():
    rows = dict(PROMO)
    rows.update({
        "A8": "Залишок січень", "B8": 100,
        "C8": "Відвантажено акційної продукції у лютому", "F8": 500,
        "H8": "витрачено акцій", "I8": 400,
        "K8": "залишок акційної продукції", "N8": 200,
    })
    block = extract_promo(sheet_of(write(rows)))
    assert block.opening == 100 and block.shipped == 500
    assert block.spent == 400 and block.closing == 200
    assert block.computed == 200


def test_promo_balance_detects_wrong_formula():
    rows = dict(PROMO)
    rows.update({
        "A8": "Залишок січень", "B8": 100,
        "C8": "Відвантажено акційної продукції у лютому", "F8": 500,
        "H8": "витрачено акцій", "I8": 400,
        "K8": "залишок акційної продукції", "N8": 1000,
    })
    block = extract_promo(sheet_of(write(rows)))
    assert block.computed == 200 and block.closing == 1000


# -- зведена ----------------------------------------------------------------

def build_consolidated() -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Акції та мотивації ТК"
    import datetime as dt
    ws["C2"] = dt.datetime(2026, 1, 1)
    heads = ["План продаж 01", "Факт продаж 01", "Бонус план 01 грн",
             "Бонус факт 01, грн", "План приріст до минулого місяця, %",
             "Факт приріст до минулого місяця, %", "Форма компенсації",
             "Затратність на мотивації з 1 пл, грн, факт"]
    ws["B3"] = "Дистрибютор"
    for i, h in enumerate(heads):
        ws.cell(row=3, column=3 + i).value = h
    ws["B4"] = "Тест Дистриб"
    for i, v in enumerate([1000, 1100, 500, 550]):
        ws.cell(row=4, column=3 + i).value = v
    ws["B5"] = "Сума"
    path = os.path.join(tempfile.mkdtemp(), "зведена.xlsx")
    wb.save(path)
    return path


def test_opening_new_month_keeps_previous_values():
    path = build_consolidated()
    cons = Consolidated(path)
    cons.ensure_month(Period(2026, 2))
    cons.write(Period(2026, 2), "sales", "Тест Дистриб",
               {"plan": 2000, "bonus_plan": 600})
    cons.save()

    again = Consolidated(path)
    jan = again.read(Period(2026, 1), "sales", "Тест Дистриб")
    feb = again.read(Period(2026, 2), "sales", "Тест Дистриб")
    assert jan["plan"] == 1000 and jan["fact"] == 1100
    assert feb["plan"] == 2000 and feb["bonus_plan"] == 600
    # приріст рахується від ФАКТУ попереднього місяця
    assert feb["growth_plan"] == "=K4/D4-100%"


def test_totals_ranges_are_repaired():
    path = build_consolidated()
    cons = Consolidated(path)
    cons.ws["C5"] = "=SUM(C4:C9)"          # діапазон заповзає за межі таблиці
    fixed = cons.refresh_totals()
    assert any("C5" in f for f in fixed)
    assert cons.ws["C5"].value == "=SUM(C4:C4)"


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-q"]))
